"""
data/daily_log.csv -> 엑셀 '일별로그' 시트로 내보내는 선택적 유틸리티 스크립트.
엑셀로 데이터를 계속 보고 싶은 경우 로컬에서 수동 실행하세요:
    python export_to_excel.py [원본 엑셀파일 경로]
저장은 "<원본파일명>_업데이트.xlsx" 로 별도 생성됩니다 (원본 파일을 직접 덮어쓰지 않음 -
엑셀에서 파일을 열어둔 상태여도 안전하게 실행할 수 있습니다).
"""
import csv
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from constants import DATA_CSV_PATH

FONT_NAME = "Arial"
BLACK = Font(name=FONT_NAME, color="000000")
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE = Font(name=FONT_NAME, size=14, bold=True, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="808080")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
KRW_FMT = '#,##0;(#,##0)"원"'
USD_FMT = "$#,##0.00"
PCT_FMT = "0.00%"

SHEET_NAME = "일별로그(자동갱신)"


def build_sheet(wb):
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    widths = [12, 13, 10, 10, 15, 16, 15, 16, 11, 11, 15, 15, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:M1")
    ws["A1"] = "일별 SpaceX 주가·환율 및 미래생명 손익 로그 (GitHub 자동수집 데이터 반영)"
    ws["A1"].font = TITLE
    ws.merge_cells("A2:M2")
    ws["A2"] = "data/daily_log.csv 를 그대로 옮긴 스냅샷입니다. 최신값은 온라인 Streamlit 대시보드를 참고하세요."
    ws["A2"].font = NOTE_FONT

    headers = ["날짜", "SpaceX종가(USD)", "환율(원)", "공모가대비등락률", "평가금액(USD)", "평가금액(KRW)",
               "투자원가(USD)", "투자원가(KRW)", "평가손익(USD)", "평가손익(KRW)", "수익률(USD)", "수익률(KRW)", "데이터출처"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    with open(DATA_CSV_PATH, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    r = 5
    for row in rows:
        vals = [row["date"], float(row["spcx_price_usd"]), float(row["fx_rate"]), float(row["pct_vs_ipo"]),
                float(row["value_usd"]), float(row["value_krw"]), float(row["cost_usd"]), float(row["cost_krw"]),
                float(row["gain_usd"]), float(row["gain_krw"]), float(row["return_usd"]), float(row["return_krw"]),
                row["source"]]
        fmts = [None, USD_FMT, "#,##0.0", PCT_FMT, USD_FMT, KRW_FMT, USD_FMT, KRW_FMT, USD_FMT, KRW_FMT, PCT_FMT, PCT_FMT, None]
        for j, (v, fmt) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = BLACK
            c.border = BORDER
            if fmt:
                c.number_format = fmt
        r += 1

    ws.freeze_panes = "A5"
    return len(rows)


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else "reference/APEX펀드_SpaceX투자_손익분석.xlsx"
    wb = openpyxl.load_workbook(src)
    n = build_sheet(wb)
    out = src.rsplit(".", 1)[0] + "_업데이트.xlsx"
    wb.save(out)
    print(f"{n}행을 '{SHEET_NAME}' 시트에 반영하여 {out} 로 저장했습니다.")


if __name__ == "__main__":
    main()
